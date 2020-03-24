# -*- coding:utf-8 -*-
# @Time :2020-03-23 18:00
# @Email :876417305@qq.com
# @Author :yanxia
# @File :do_excel.PY
import openpyxl
from common import http_request
from common import file_adress
from common.read_config import config
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.method=None
        self.url=None
        self.data=None
        self.expected=None
        self.actual=None
        self.result=None
        self.sql=None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.wb=openpyxl.load_workbook(file_name)
        self.shee_name=sheet_name
        self.sheet=self.wb[sheet_name]
    def get_cases(self):
        mode=eval(config.get("CASE_MODE","mode"))
        test_data=[]
        for key in mode:
            self.sheet=self.wb[key]
            if mode[key]=='all':
                for i in range(2,self.sheet.max_row+1):
                    case=Case()
                    case.case_id=self.sheet.cell(row=i,column=1).value
                    case.title = self.sheet.cell(row=i, column=2).value
                    case.method = self.sheet.cell(row=i, column=3).value
                    case.url = self.sheet.cell(row=i, column=4).value
                    case.data = self.sheet.cell(row=i, column=5).value
                    case.expected = self.sheet.cell(row=i, column=6).value
                    case.sql = self.sheet.cell(row=i, column=9).value
                    case.sheet_name=key
                    test_data.append(case)
            else:
                for case_id in mode[key]:
                    case = Case()
                    case.case_id = self.sheet.cell(row=case_id+1, column=1).value
                    case.title = self.sheet.cell(row=case_id+1, column=2).value
                    case.method = self.sheet.cell(row=case_id+1, column=3).value
                    case.url = self.sheet.cell(row=case_id+1, column=4).value
                    case.data = self.sheet.cell(row=case_id+1, column=5).value
                    case.expected = self.sheet.cell(row=case_id+1, column=6).value
                    case.sql = self.sheet.cell(row=case_id+1, column=9).value
                    case.sheet_name = key
                    test_data.append(case)
        self.wb.close()
        return test_data
    def write_result(self,row,actual,result):
        sheet=self.wb[self.shee_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.wb.save(filename=self.file_name)
        self.wb.close()
if __name__ == '__main__':
    do_excel=DoExcel(file_adress.data_path,sheet_name='login')
    cases=do_excel.get_cases()
    http_request=http_request.HttpRequest()
    for i in cases:
        print(i.__dict__)
    res=http_request.request(i.method,i.url,i.data)
    actual=res.text
    if i.expected==actual:
        do_excel.write_result(i.case_id+1,actual,'FASS')
    else:
        do_excel.write_result(i.case_id + 1, actual, 'Fail')





